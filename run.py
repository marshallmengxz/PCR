#!/usr/bin/env python
from xlrd import open_workbook
from openpyxl import load_workbook
import tkinter as tk
from PIL import Image, ImageTk
import pytesseract


def bossID(bossname):
    if bossname=='双足飞龙':
        return(1)
    elif bossname=='野性狮鹫':
        return(2)
    elif bossname=='兽人头目':
        return(3)
    elif bossname=='灵魂角鹿':
        return(4)
    elif bossname=='弥诺陶洛斯':
        return (5)




def read_data(image_path):
    imageObject = Image.open(image_path)
    x_start = int(x1.get('1.0','end-1c'))
    x_end = int(x2.get('1.0','end-1c'))
    y_start = int(y1.get('1.0','end-1c'))
    y_end =  int(y2.get('1.0','end-1c'))
    cropped = imageObject.crop((x_start, y_start, x_end, y_end))


    text = pytesseract.image_to_string(cropped, lang='chi_sim+eng+chi_tra', config='--psm 1')
    # print(text)
    text = text.replace(' ', '')
    text = text.replace('\n', '')
    print(text)
    playerid_set = []
    damage_set = []
    bossid_set = []
    while (text and len(text) > 5):
        playerid = text[0:text.find('对')]
        damage = text[text.find('造成了') + 3:text.find('伤害')]
        bossid = bossID(text[text.find('对') + 1:text.find('造成了')])
        playerid_set.append(playerid)
        damage_set.append(damage)
        bossid_set.append(bossid)
        text = text[text.find('伤害') + 2:]
    return cropped, playerid_set, damage_set, bossid_set



def auto_read():
    path.set(path_entry.get('1.0','end-1c'))
    cropped, playerid_set, damage_set, bossid_set = read_data(path.get())
    resized = cropped.resize((150, 215), Image.ANTIALIAS)
    image_file = ImageTk.PhotoImage(resized)
    image = tk.Label(window,image=image_file)
    image.img=image_file
    image.pack()
    image.place(x=300, y=300)
    id1.set(playerid_set[0])
    id2.set(playerid_set[1])
    id3.set(playerid_set[2])
    id4.set(playerid_set[3])
    bi1.set(bossid_set[0])
    bi2.set(bossid_set[1])
    bi3.set(bossid_set[2])
    bi4.set(bossid_set[3])
    d1.set(damage_set[0])
    d2.set(damage_set[1])
    d3.set(damage_set[2])
    d4.set(damage_set[3])

def auto_write():
    loginfo.delete('1.0','end')
    savePath = save_path.get('1.0','end-1c')
    date = int(day.get('1.0','end-1c'))
    # 读取excel文件
    myexcel = open_workbook(savePath)
    table = myexcel.sheets()[0]

    rows = table.nrows
    columns = table.ncols
    table_content = []
    for i in range(columns):
        table_content.append([])
    print(rows, columns)
    for j in range(columns):
        for i in range(rows):
            table_content[j].append(table.cell_value(i, j))  # 返回单元格中的数据

    # ID 列表
    id_set = table_content[0]
    # 修改excel文件
    data = load_workbook(savePath)  # 可读可写
    present_round=1
    ws = data['Sheet1']
    playerid_set = []
    playerid_set.append(id1.get())
    playerid_set.append(id2.get())
    playerid_set.append(id3.get())
    playerid_set.append(id4.get())
    damage_set = []
    damage_set.append(d1.get())
    damage_set.append(d2.get())
    damage_set.append(d3.get())
    damage_set.append(d4.get())
    bossid_set = []
    bossid_set.append(bi1.get())
    bossid_set.append(bi2.get())
    bossid_set.append(bi3.get())
    bossid_set.append(bi4.get())
    turn_set = []
    turn_set.append(turn1.get('1.0','end-1c'))
    turn_set.append(turn2.get('1.0', 'end-1c'))
    turn_set.append(turn3.get('1.0', 'end-1c'))
    turn_set.append(turn4.get('1.0', 'end-1c'))

    for playerid, bossid, damage, turn in zip(playerid_set, bossid_set, damage_set, turn_set):
        if playerid in id_set:
            print(playerid, bossid, damage, turn)
            id_index = id_set.index(playerid)
            date_index = 2 + (date - 1) * 3
            if (table_content[date_index][id_index] and [ord(c) for c in str(table_content[date_index + 1][id_index])] != [
                ord(c) for c in damage]):
                id_index = id_index + 1
                if (table_content[date_index][id_index] and [ord(c) for c in table_content[date_index + 1][id_index]] != [
                    ord(c) for c in damage]):
                    id_index = id_index + 1

            ws.cell(row=id_index + 1, column=date_index + 1, value=bossid)
            table_content[date_index][id_index] = bossid
            ws.cell(row=id_index + 1, column=date_index + 2, value=damage)
            ws.cell(row=id_index + 1, column=date_index, value=turn)
        else:
            loginfo.insert('end',playerid+' is not in the ID list\n\n')
    data.save(savePath)

def auto_info():
    savePath=save_path.get('1.0','end-1c')
    myexcel = open_workbook(savePath)
    table = myexcel.sheets()[0]
    rows = table.nrows
    id_set = []
    for i in range(rows):
        id_set.append(table.cell_value(i, 0))  # 返回单元格中的数据
    id_set = filter(None, id_set)
    for id in id_set:
        box1.insert('end',id)



window = tk.Tk()

window.title('My window')
window.geometry('700x700')
cropped=0
path = tk.StringVar()    # 将label标签的内容设置为字符类型，用var来接收hit_me函数的传出内容用以显示在标签上
id1 = tk.StringVar()
id2 = tk.StringVar()
id3 = tk.StringVar()
id4 = tk.StringVar()
bi1 = tk.StringVar()
bi2 = tk.StringVar()
bi3 = tk.StringVar()
bi4 = tk.StringVar()
d1 = tk.StringVar()
d2 = tk.StringVar()
d3 = tk.StringVar()
d4 = tk.StringVar()
log='Successfully added 4 records!'



path_entry = tk.Text(window, font=('Arial', 14))
path_entry.insert('insert','/Users/mengzixia/Desktop/damage/8.jpg')
path_entry.pack()
path_entry.place(height=30,width=330,x=150,y=30)

save_path = tk.Text(window, font=('Arial', 14))
save_path.insert('insert','/Users/mengzixia/Downloads/工会战报刀表.xlsx')
save_path.pack()
save_path.place(height=30,width=330,x=300,y=600)

path_label = tk.Label(window,text='图片地址',fg='black', font=('Arial', 12))
path_label.pack()
path_label.place(height=30,width=50,x=80,y=30)

label1 = tk.Label(window,text='截取范围',fg='black', font=('Arial', 12))
label1.pack()
label1.place(height=30,width=150,x=30,y=60)

label2 = tk.Label(window,text='x',fg='black', font=('Arial', 12))
label2.pack()
label2.place(height=30,width=10,x=150,y=60)
x1=tk.Text(window, height=2, spacing1=2)
x1.insert('insert','1493')
x1.pack()
x1.place(height=30,width=50,x=200,y=60)
x2=tk.Text(window, height=2, spacing1=2)
x2.insert('insert','1750')
x2.pack()
x2.place(height=30,width=50,x=320,y=60)




y1 = tk.Text(window, height=2, spacing1=2)
y1.insert('insert','220')
y1.pack()
y1.place(height=30, width=50, x=200, y=90,)
y2 = tk.Text(window, height=2, spacing1=2)
y2.insert('insert','700')
y2.pack()
y2.place(height=30,width=50,x=320,y=90)

label3 = tk.Label(window, text='~', fg='black', font=('Arial', 12))
label3.pack()
label3.place(height=30, width=50, x=250, y=60)

label3 = tk.Label(window, text='~',fg='black', font=('Arial', 12))
label3.pack()
label3.place(height=30, width=50, x=250, y=90)

label4 = tk.Label(window, text='y', fg='black', font=('Arial', 12))
label4.pack()
label4.place(height=30,width=10,x=150,y=90)

label5 = tk.Label(window,text='Player_id',fg='black', font=('Arial', 12))
label5.pack()
label5.place(height=30,width=50,x=150,y=120)

label6 = tk.Label(window,text='Boss_id',fg='black', font=('Arial', 12))
label6.pack()
label6.place(height=30,width=50,x=230,y=120)

label7 = tk.Label(window,text='Damage',fg='black', font=('Arial', 12))
label7.pack()
label7.place(height=30, width=50, x=300, y=120)

label8 = tk.Label(window, text='轮次', fg='black', font=('Arial', 12))
label8.pack()
label8.place(height=30, width=50, x=80, y=120)

label9 = tk.Label(window, text='识别区域预览', fg='black', font=('Arial', 12))
label9.pack()
label9.place(height=30, width=80, x=300, y=270)

label10 = tk.Label(window, text='工会成员', fg='black', font=('Arial', 12))
label10.pack()
label10.place(height=30, width=80, x=70, y=270)

label11 = tk.Label(window, text='工会成员文件路径', fg='black', font=('Arial', 12))
label11.pack()
label11.place(height=30, width=120, x=290, y=570)

label12 = tk.Label(window,text='会战日', fg='black', font=('Arial', 12))
label12.pack()
label12.place(height=30, width=120, x=265, y=540)

loginfo = tk.Text(window, fg='black', font=('Arial', 15),spacing1=2)
loginfo.pack()
loginfo.place(height=200, width=150, x=500, y=50)

b_info = tk.Button(window, text='获取成员信息', font=('Arial', 12), width=15, height=1, command=auto_info)
b_info.pack()
b_info.place(x=150,y=275)

sb = tk.Scrollbar(window)
box1 = tk.Listbox(window,height=20,selectmode='single',yscrollcommand=sb.set)
box1.pack(side='left',fill='both')
sb.config(command=box1.yview)
sb.pack(side='right',fill='y')

box1.place(x=70,y=300)


turn1 = tk.Text(window, height=2, spacing1=2)
turn1.insert('insert','1')
turn1.pack()
turn1.place(height=30,width=50,x=90,y=150)

turn2 = tk.Text(window, height=2, spacing1=2)
turn2.insert('insert','1')
turn2.pack()
turn2.place(height=30,width=50,x=90,y=180)

turn3 = tk.Text(window, height=2, spacing1=2)
turn3.insert('insert','1')
turn3.pack()
turn3.place(height=30,width=50,x=90,y=210)

turn4 = tk.Text(window, height=2, spacing1=2)
turn4.insert('insert','1')
turn4.pack()
turn4.place(height=30,width=50,x=90,y=240)

day = tk.Text(window,height=2,spacing1=2)
day.insert('insert','1')
day.pack()
day.place(height=30,width=50,x=350,y=540)

# id_widge1 = tk.Text(window, height=2, spacing1=2)
# id_widge1.delete('1.0','end')
# id_widge1.insert('insert',id1.get())

id_widge1 = tk.Entry(window,textvariable=id1, font=('Arial', 9))
id_widge1.pack()
id_widge1.place(height=30, width=100, x=120, y=150)

id_widge2 = tk.Entry(window,textvariable=id2, font=('Arial', 9))
id_widge2.pack()
id_widge2.place(height=30, width=100, x=120, y=180)

id_widge3 = tk.Entry(window,textvariable=id3, font=('Arial', 9))
id_widge3.pack()
id_widge3.place(height=30, width=100, x=120, y=210)

id_widge4 = tk.Entry(window,textvariable=id4, font=('Arial', 9))
id_widge4.pack()
id_widge4.place(height=30, width=100, x=120, y=240)

boss_widge1 = tk.Entry(window, textvariable=bi1, font=('Arial', 9))
boss_widge1.pack()
boss_widge1.place(height=30, width=30, x=240, y=150)

boss_widge2 = tk.Entry(window, textvariable=bi2, font=('Arial', 9))
boss_widge2.pack()
boss_widge2.place(height=30, width=30, x=240, y=180)

boss_widge3 = tk.Entry(window, textvariable=bi3, font=('Arial', 9))
boss_widge3.pack()
boss_widge3.place(height=30, width=30, x=240, y=210)

boss_widge4 = tk.Entry(window, textvariable=bi4, font=('Arial', 9))
boss_widge4.pack()
boss_widge4.place(height=30, width=30, x=240, y=240)

damage_widge1 = tk.Entry(window, textvariable=d1, font=('Arial', 9))
damage_widge1.pack()
damage_widge1.place(height=30, width=50, x=300, y=150)

damage_widge2 = tk.Entry(window, textvariable=d2, font=('Arial', 9))
damage_widge2.pack()
damage_widge2.place(height=30, width=50, x=300, y=180)

damage_widge3 = tk.Entry(window, textvariable=d3, font=('Arial', 9))
damage_widge3.pack()
damage_widge3.place(height=30, width=50, x=300, y=210)

damage_widge4 = tk.Entry(window, textvariable=d4, font=('Arial', 9))
damage_widge4.pack()
damage_widge4.place(height=30, width=50, x=300, y=240)

b_read = tk.Button(window, text='获取伤害数据', font=('Arial', 12), width=10, height=1, command=auto_read)
b_read.pack()
b_read.place(height=30,width=100,x=380,y=75)

b_write = tk.Button(window, text='录入', font=('Arial', 12), width=10, height=1, command=auto_write)
b_write.pack()
b_write.place(height=150,width=100,x=380,y=120)



window.mainloop()

